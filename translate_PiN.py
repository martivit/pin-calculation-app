import numpy as np
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.cell import MergedCell  # Import MergedCell
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from io import BytesIO


translation_dict = {
    'Girls (5-17 y.o.)': 'Filles (5-17 ans)',
    'Boys (5-17 y.o.)': 'Garçons (5-17 ans)',
    'TOTAL (5-17 y.o.)': 'TOTAL (5-17 ans)'
}



def translate_excel_sheets_with_formatting(excel_data):
    # Load the workbook from the BytesIO object
    workbook = openpyxl.load_workbook(excel_data)

    # Iterate over each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate over each cell in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Replace strings based on the translation dictionary
                    for key, value in translation_dict.items():
                        if key in cell.value:
                            cell.value = cell.value.replace(key, value)
    
    # Save the translated workbook back into a BytesIO object
    translated_output = BytesIO()
    workbook.save(translated_output)
    translated_output.seek(0)
    
    return translated_output
