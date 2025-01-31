import numpy as np
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.cell import MergedCell  # Import MergedCell
from openpyxl.utils.dataframe import dataframe_to_rows


int_2 = '2.0'
int_3 = '3.0'
int_4 = '4.0'
int_5 = '5.0'
label_perc2 = '% severity levels 1-2'
label_perc3 = '% severity level 3'
label_perc4 = '% severity level 4'
label_perc5 = '% severity level 5'
label_tot2 = '# severity levels 1-2'
label_tot3 = '# severity level 3'
label_tot4 = '# severity level 4'
label_tot5 = '# severity level 5'
label_perc_tot = '% Tot PiN (severity levels 3-5)'
label_tot = '# Tot PiN (severity levels 3-5)'
label_admin_severity = 'Area severity'
label_tot_population = 'TotN'

int_acc = 'access'
int_agg= 'aggravating circumstances'
int_lc = 'learning condition'
int_penv = 'protected environment'
int_out = 'Not in need'
label_perc_acc = '% Access'
label_perc_agg= '% Aggravating circumstances'
label_perc_lc = '% Learning conditions'
label_perc_penv = '% Protected environment'
label_perc_out = '% Not in need'
label_tot_acc = '# Access'
label_tot_agg= '# Aggravating circumstances'
label_tot_lc = '# Learning conditions'
label_tot_penv = '# Protected environment'
label_tot_out = '# Not in need'
label_dimension_perc_tot = '% Tot in PiN Dimensions'
label_dimension_tot = '# Tot in PiN Dimensions'
label_dimension_tot_population = 'TotN'


# Define the colors
colors = {
    "light_beige": "FFF2CC",
    "light_orange": "F4B183",
    "dark_orange": "ED7D31",
    "darker_orange": "C65911",
    "light_blue": "DDEBF7",
    "light_pink": "b3b389",
    "light_yellow": "ffffc5",
    "white": "FFFFFF",
    "bluepin": "004bb4",
    'gray': 'e0e0e0',
    'stratagray': 'F0F0F0'
}
# Define the columns to color
color_mapping = {
    label_perc2: colors["light_beige"],
    label_tot2: colors["light_beige"],
    label_perc3: colors["light_orange"],
    label_tot3: colors["light_orange"],
    label_perc4: colors["dark_orange"],
    label_tot4: colors["dark_orange"],
    label_perc5: colors["darker_orange"],
    label_tot5: colors["darker_orange"],
    label_perc_tot: colors["light_blue"],
    label_admin_severity: colors["light_blue"],
    label_tot: colors["light_blue"]
}
# Define the colors
colors_dimension = {
    "light_beige": "ebecc7",
    "light_orange": "c7ebec",
    "dark_orange": "c7d9ec",
    "darker_orange": "c7ecdb",
    'darker2_orange':'b3d3d4',
    "light_blue": "DDEBF7",
    "light_pink": "b3b389",
    "light_yellow": "ffffc5",
    "white": "FFFFFF",
    "bluepin": "004bb4",
    'gray': 'e0e0e0',
    'stratagray': 'F0F0F0'
}
# Define the columns to color
color_mapping_dimension = {
    label_perc_out: colors_dimension["light_beige"],
    label_tot_out: colors_dimension["light_beige"],
    label_perc_acc: colors_dimension["light_orange"],
    label_tot_acc: colors_dimension["light_orange"],
    label_perc_agg: colors_dimension["dark_orange"],
    label_tot_agg: colors_dimension["dark_orange"],
    label_perc_lc: colors_dimension["darker_orange"],
    label_tot_lc: colors_dimension["darker_orange"],
    label_perc_penv: colors_dimension["darker2_orange"],
    label_tot_penv: colors_dimension["darker2_orange"],
    label_dimension_perc_tot: colors_dimension["light_blue"],
    label_dimension_tot: colors_dimension["light_blue"]
}

alignment_columns = list(color_mapping.keys())
def apply_final_formatting(country_name, workbook, overview_df, small_overview_df, admin_var, selected_language= 'English'):


    label_perc2 = '% severity levels 1-2'
    label_perc3 = '% severity level 3'
    label_perc4 = '% severity level 4'
    label_perc5 = '% severity level 5'
    label_tot2 = '# severity levels 1-2'
    label_tot3 = '# severity level 3'
    label_tot4 = '# severity level 4'
    label_tot5 = '# severity level 5'
    label_perc_tot = '% Tot PiN (severity levels 3-5)'
    label_tot = '# Tot PiN (severity levels 3-5)'
    label_admin_severity = 'Area severity'
    label_tot_population = 'TotN'

    color_mapping = {
        label_perc2: colors["light_beige"],
        label_tot2: colors["light_beige"],
        label_perc3: colors["light_orange"],
        label_tot3: colors["light_orange"],
        label_perc4: colors["dark_orange"],
        label_tot4: colors["dark_orange"],
        label_perc5: colors["darker_orange"],
        label_tot5: colors["darker_orange"],
        label_perc_tot: colors["light_blue"],
        label_admin_severity: colors["light_blue"],
        label_tot: colors["light_blue"]
    }

    tot_5_17_label = 'TOTAL (5-17 y.o.)'
    girl_5_17_label = 'Girls (5-17 y.o.)'
    boy_5_17_label = 'Boys (5-17 y.o.)'
    ece_5yo_label = 'ECE (5 y.o.)'
    age_label = ' (5-17 y.o.)'
    if country_name == 'Afghanistan':
        tot_5_17_label = 'TOTAL (6-17 y.o.)'
        girl_5_17_label = 'Girls (6-17 y.o.)'
        boy_5_17_label = 'Boys (6-17 y.o.)'
        ece_5yo_label = 'ECE (6 y.o.)'
        age_label = ' (6-17 y.o.)'

    if selected_language == "French":
        tot_5_17_label= 'TOTAL (5-17 ans)'
        girl_5_17_label= 'Filles (5-17 ans)'
        boy_5_17_label='Garcons (5-17 ans)'
        ece_5yo_label= 'Éducation préscolaire (5 ans)'
        age_label = ' (5-17 ans)'
        label_perc2= '% niveaux de sévérité 1-2'
        label_perc3= '% niveau de sévérité 3'
        label_perc4= '% niveau de sévérité 4'
        label_perc5= '% niveau de sévérité 5'
        label_tot2= '# niveaux de sévérité 1-2'
        label_tot3= '# niveau de sévérité 3'
        label_tot4= '# niveau de sévérité 4'
        label_tot5= '# niveau de sévérité 5'
        label_perc_tot= "% Tot PiN (niveaux de sévérité 3-5)"
        label_tot= "# Tot PiN (niveaux de sévérité 3-5)"
        label_admin_severity= 'Sévérité de la zone'
        label_tot_population= 'Population totale'
        color_mapping = {
            label_perc2: colors["light_beige"],
            label_tot2: colors["light_beige"],
            label_perc3: colors["light_orange"],
            label_tot3: colors["light_orange"],
            label_perc4: colors["dark_orange"],
            label_tot4: colors["dark_orange"],
            label_perc5: colors["darker_orange"],
            label_tot5: colors["darker_orange"],
            label_perc_tot: colors["light_blue"],
            label_admin_severity: colors["light_blue"],
            label_tot: colors["light_blue"]
        }

    for ws in workbook.worksheets:
        if ws.title == "PiN TOTAL":
            # Clear existing content in the worksheet
            ws.delete_rows(1, ws.max_row)

            # Add empty rows at the top and adjust columns
            ws.insert_rows(1, 4)
            ws.insert_cols(1, 2)

            # Write small_overview_df on the left (columns C-E)
            for r_idx, row in enumerate(dataframe_to_rows(small_overview_df, index=False, header=True), start=5):
                for c_idx, value in enumerate(row, start=3):  # Start writing in column C
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Write overview_df on the right (columns H onwards)
            for r_idx, row in enumerate(dataframe_to_rows(overview_df, index=False, header=True), start=5):
                for c_idx, value in enumerate(row, start=8):  # Start writing in column H
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Merge the title cell across all filled columns and ensure it is centered
            max_col = ws.max_column
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=max_col)
            title_cell = ws.cell(row=1, column=3)
            title_cell.value = "PiN TOTAL"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Set header formatting (bold, orange fill)
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=colors["light_orange"], end_color=colors["light_orange"], fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply color formatting and borders based on the content of cells
            for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
                strata_value = row[2].value  # Adjusting for zero-indexing; column C is the 'Strata' column
                
                if strata_value == tot_5_17_label:
                    fill_color = colors["bluepin"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.font = Font(color=colors["white"], bold=True)  # Set text color to white and bold
                elif strata_value in [girl_5_17_label, boy_5_17_label]:
                    fill_color = colors["gray"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                elif strata_value in ['Female', 'Male', ece_5yo_label, 'Children with disability', 'Enfants en situation de handicap', 'Filles', 'Garcons']:
                    fill_color = colors["stratagray"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                else:
                    fill_color = colors["light_blue"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Set column widths based on content
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if cell.value and not isinstance(cell, MergedCell):
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                                column_letter = cell.column_letter  # Get the column letter
                            except Exception as e:
                                print(f"Error: {e}, Cell: {cell}")

                    if column_letter:
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column_letter].width = adjusted_width

            # Set the first two columns (A and B) to white background
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")

            # Apply specific formatting to columns F and G
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=7):
                for cell in row:
                    # Example formatting, adjust as needed
                    cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")
                    cell.alignment = Alignment(horizontal='right', vertical='center')

               # Set background for empty cells and apply black borders to cells with values
            thin_border = Border(
                left=Side(style='thin', color="000000"),
                right=Side(style='thin', color="000000"),
                top=Side(style='thin', color="000000"),
                bottom=Side(style='thin', color="000000")
            )

            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=ws.max_column):  # Adjust columns as needed
                for cell in row:
                    if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):  
                        cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")
                    else:  # If the cell has a value, apply black border
                        cell.border = thin_border

        if ws.title != "PiN TOTAL":
            ws.insert_rows(1, 4)
            # Add empty columns to the left
            ws.insert_cols(1, 4)
            # Add the sheet name as a title in the first row
            title = ws.title
            title += age_label
            max_col = ws.max_column
            ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=max_col)
            title_cell = ws.cell(row=1, column=5)
            title_cell.value = title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Set column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if cell.value and not isinstance(cell, MergedCell):
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                            column_letter = cell.column_letter  # Get the column letter
                        except Exception as e:
                            print(f"Error: {e}, Cell: {cell}")

                if column_letter:
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Bold specific columns
            for row in ws.iter_rows(min_row=6, max_col=ws.max_column, max_row=ws.max_row):  # Start from the data row
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value  # Row 5 contains the headers
                    if col_name in [label_perc_tot, label_admin_severity, label_tot, admin_var]:
                        cell.font = Font(bold=True)  # Apply bold font
        
            # Apply formatting to the headers
            for cell in ws[5]:  # Header row is now the 5th row
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply color to specific columns and make borders visible
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value
                    # Apply color based on mapping
                    if col_name in color_mapping:
                        cell.fill = PatternFill(start_color=color_mapping[col_name], end_color=color_mapping[col_name], fill_type="solid", patternType='solid')
                    # Apply alignment
                    if col_name in alignment_columns:
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                    # Apply border, but skip the first four columns
                    if cell.column > 4:
                        if row[0].row == 5:  # Bold top border for header
                            cell.border = Border(
                                top=Side(style="medium"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[0]:  # Bold left border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="medium"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[-1]:  # Bold right border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="medium"),
                                bottom=Side(style="thin"),
                            )
                        elif row[0].row == ws.max_row:  # Bold bottom border for last row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="medium"),
                            )
                        else:
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                    else:
                        cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")  # White background for the first four columns
                        cell.border = None  # No border for the first four columns



    return workbook


# Function to create output with final formatting
def create_output(country_label, dataframes, overview_df, small_overview_df, overview_sheet_name, admin_var, ocha=True, tot_severity=None, selected_language='English', parameters=None):
    country_name = country_label.split('__')[0]  # Extract the part before the "__"

    label_overall_severity = 'Overall PiN and severity'
    if selected_language == "French":
        label_overall_severity = 'PiN total par admin'
    output = BytesIO()
    with pd.ExcelWriter(output) as writer:
        # Only write the overview sheet if ocha is True
        if ocha:
            overview_df.to_excel(writer, sheet_name=overview_sheet_name, index=False)

        # Write the tot_severity sheet if it is provided
        if tot_severity is not None:
            tot_severity.to_excel(writer, sheet_name=label_overall_severity, index=False)

        # Write the category sheets
        for category, df in dataframes.items():
            sheet_name = f"{overview_sheet_name.split()[0]} -- {category}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        if parameters:
            if selected_language == "English":
                parameters_df = pd.DataFrame(
                    [{"Category": "General Information", "Key": key, "Value": value}
                    for key, value in parameters["general_info"].items()] +
                    [{"Category": "MSNA Indicators per PiN dimension", "Key": f"{key} - {sub_key}", "Value": sub_value}
                    for key, sub_dict in parameters["msna_indicators_per_PiN_dimension"].items()
                    for sub_key, sub_value in (sub_dict.items() if isinstance(sub_dict, dict) else [(key, sub_dict)])] +
                    [{"Category": "Severity Classification", "Key": f"{key} - {sub_key}", "Value": sub_value}
                    for key, sub_dict in parameters["severity_classification"].items()
                    for sub_key, sub_value in sub_dict.items()] +
                    [{"Category": "HNO Unit of analysis", "Key": key, "Value": value}
                    for key, value in parameters["admin_unit"].items()] +
                    [{"Category": "School Cycles", "Key": key, "Value": value}
                    for key, value in parameters["school_cycles"].items()]
                )
                parameters_df.to_excel(writer, sheet_name="Parameters Used", index=False)
            elif selected_language == "French":
                parameters_df = pd.DataFrame(
                    [{"Catégorie": "Informations générales", "Clé": key, "Valeur": value}
                    for key, value in parameters["informations_generales"].items()] +
                    [{"Catégorie": "Indicateurs MSNA par dimension du PiN", "Clé": f"{key} - {sub_key}", "Valeur": sub_value}
                    for key, sub_dict in parameters["indicateurs_msna_par_dimension"].items()
                    for sub_key, sub_value in (sub_dict.items() if isinstance(sub_dict, dict) else [(key, sub_dict)])] +
                    [{"Catégorie": "Classification de la sévérité", "Clé": f"{key} - {sub_key}", "Valeur": sub_value}
                    for key, sub_dict in parameters["classification_de_sévérité"].items()
                    for sub_key, sub_value in sub_dict.items()] +
                    [{"Catégorie": "Unité d’analyse HNO", "Clé": key, "Valeur": value}
                    for key, value in parameters["unité_administrative"].items()] +
                    [{"Catégorie": "Cycles scolaires", "Clé": key, "Valeur": value}
                    for key, value in parameters["cycles_scolaires"].items()]
                )
                parameters_df.to_excel(writer, sheet_name="Paramètres Utilisés", index=False)


    output.seek(0)
    workbook = load_workbook(output)

    # Apply the final formatting to the workbook
    workbook = apply_final_formatting(country_name,workbook, overview_df, small_overview_df, admin_var, selected_language=selected_language)
    
    formatted_output = BytesIO()
    workbook.save(formatted_output)
    formatted_output.seek(0)

    return formatted_output


def create_indicator_output(country_label, indicator_dataframes, admin_var, selected_language='English'):
    """
    Creates an Excel file for indicator-based data, applying formatting.

    Parameters:
    - country_label (str): The name of the country (used in the file name).
    - indicator_dataframes (dict): Dictionary of DataFrames categorized by indicator.
    - admin_var (str): The administrative variable used in the dataset.
    - selected_language (str, default='English'): Language setting for headers.

    Returns:
    - BytesIO: The formatted Excel file as an in-memory object.
    """
    country_name = country_label.split('__')[0]  # Extract country name

    # File output buffer
    output = BytesIO()
    
    with pd.ExcelWriter(output) as writer:
        # Modify column names BEFORE writing them to the Excel file
        modified_dataframes = {}

        for category, df in indicator_dataframes.items():
            # Rename columns: Add (% of children) after ":" unless they have (ToT # children)
            new_columns = {}
            for col in df.columns:
                if ":" in col and "(ToT # children)" not in col:
                    new_columns[col] = col.replace(":", ": (% of children)", 1)

            df = df.rename(columns=new_columns)
            modified_dataframes[category] = df

            # Write to Excel, ensuring sheet names stay within limits
            df.to_excel(writer, sheet_name=category[:30], index=False)

    # Load the workbook for formatting
    output.seek(0)
    workbook = load_workbook(output)

    for ws in workbook.worksheets:
        ws.insert_rows(1, 4)  # Add empty rows at the top
        ws.insert_cols(1, 4)  # Add empty columns on the left

        # **Increase header row thickness more**
        ws.row_dimensions[5].height = 50  # Make row even thicker

        # Title formatting
        title = ws.title
        max_col = ws.max_column
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=5)
        title_cell.value = f"Children (5–17 years old) classified by severity and indicators"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Extract headers from row 5
        headers = [ws.cell(row=5, column=col).value for col in range(1, ws.max_column + 1)]
        
        # **Increase Column Widths Based on Content & Enable Wrap Text**
        for col_idx, col_name in enumerate(headers, start=1):
            max_length = max((len(str(ws.cell(row=row_idx, column=col_idx).value)) for row_idx in range(5, ws.max_row + 1)), default=10)
            adjusted_width = max(15, min(max_length + 2, 25))  # Ensure minimum width but not too wide
            ws.column_dimensions[ws.cell(row=5, column=col_idx).column_letter].width = adjusted_width

            # Apply **wrap text** to headers
            header_cell = ws.cell(row=5, column=col_idx)
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_cell.font = Font(bold=True, size=10)  # **Reduce font size**

        # Apply color to specific columns and make borders visible
        for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                col_index = cell.column  # Get column index
                col_name = headers[col_index - 1] if col_index - 1 < len(headers) else None  # Prevent index error

                if col_name and isinstance(col_name, str):  # Ensure col_name is valid
                    # Apply color based on severity level
                    if "severity level 3" in col_name:
                        cell.fill = PatternFill(start_color=colors["light_orange"], end_color=colors["light_orange"], fill_type="solid")
                    elif "severity level 4" in col_name:
                        cell.fill = PatternFill(start_color=colors["dark_orange"], end_color=colors["dark_orange"], fill_type="solid")
                    elif "severity level 5" in col_name:
                        cell.fill = PatternFill(start_color=colors["darker_orange"], end_color=colors["darker_orange"], fill_type="solid")

                # Apply wrap text and reduce font size for all data cells
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                cell.font = Font(size=10)  # Reduce font size for better readability

                # Apply border formatting
                if col_index > 4:
                    if row[0].row == 5:  # Bold top border for header
                        cell.border = Border(top=Side(style="medium"), left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))
                    elif row[0].row == ws.max_row:  # Bold bottom border for last row
                        cell.border = Border(top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="medium"))
                    else:
                        cell.border = Border(top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))

    # Save formatted workbook
    formatted_output = BytesIO()
    workbook.save(formatted_output)
    formatted_output.seek(0)

    return formatted_output