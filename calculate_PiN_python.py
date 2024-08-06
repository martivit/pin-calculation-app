import pandas as pd
import fuzzywuzzy
from fuzzywuzzy import process
import numpy as np
import datetime
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.cell import MergedCell  # Import MergedCell



##--------------------------------------------------------------------------------------------
def calculate_age_correction(start_month, collection_month):
    # Create a dictionary to map the first three letters of month names to their numeric equivalents
    month_lookup = {datetime.date(2000, i, 1).strftime('%b').lower(): i for i in range(1, 13)}
    
    # Convert month names to their numeric equivalents using the predefined lookup
    start_month_num = month_lookup[start_month.strip()[:3].lower()]
    
    # Adjust the start month number for a school year starting in the previous calendar year
    adjusted_start_month_num = start_month_num - 12 if start_month_num > 6 else start_month_num
    
    # Determine if the age correction should be applied based on the month difference
    age_correction = (collection_month - adjusted_start_month_num) > 6
    return age_correction

##--------------------------------------------------------------------------------------------
def find_matching_choices(choices_df, barriers_list):
    # List to hold the results
    results = []
    
    # Iterate over each barrier in the list
    for barrier in barriers_list:
        # Filter choices where 'label::english' matches the current barrier
        matched_choices = choices_df[choices_df['label::english'] == barrier]
        
        # For each matched choice, create an entry in the results list
        for _, choice in matched_choices.iterrows():
            result_entry = {'name': choice['name'], 'label': barrier}
            results.append(result_entry)
    
    return results
##--------------------------------------------------------------------------------------------
def assign_school_cycle(edu_age_corrected, single_cycle=False, lower_primary_start_var=6, lower_primary_end_var=13, upper_primary_end_var=None):
    if single_cycle:
        # If single cycle is True, handle as a primary to secondary without upper primary
        if lower_primary_start_var <= edu_age_corrected <= lower_primary_end_var:
            return 'primary'
        elif lower_primary_end_var + 1 <= edu_age_corrected <= 18:
            return 'secondary'
        elif edu_age_corrected == 5: 
            return 'ECE'
        else:
            return 'out of school range'
    else:
        # If single cycle is False, handle lower primary, upper primary, and secondary phases
        if lower_primary_start_var <= edu_age_corrected <= lower_primary_end_var:
            return 'lower primary'
        elif upper_primary_end_var and lower_primary_end_var + 1 <= edu_age_corrected <= upper_primary_end_var:
            return 'upper primary'
        elif upper_primary_end_var and upper_primary_end_var + 1 <= edu_age_corrected <= 18:
            return 'secondary'
        elif edu_age_corrected == 5: 
            return 'ECE'
        else:
            return 'out of school range'
        
##--------------------------------------------------------------------------------------------
def calculate_severity(access, barrier, armed_disruption, idp_disruption, teacher_disruption, names_severity_4, names_severity_5):
    # Helper function to safely normalize string inputs
    def normalize(input_string):
        if isinstance(input_string, str):
            return input_string.lower()
        return ""  # Default to empty string if input is not a string
    
    # Normalize the input to handle different cases and languages
    normalized_access = normalize(access)
    normalized_armed_disruption = normalize(armed_disruption)
    normalized_idp_disruption = normalize(idp_disruption)
    normalized_teacher_disruption = normalize(teacher_disruption)

    # Normalize to handle English and French variations of "yes" and "no"
    yes_answers = ['yes', 'oui']
    no_answers = ['no', 'non']
    

    if normalized_access in no_answers:
        if barrier in names_severity_5:
            return 5
        elif barrier in names_severity_4:
            return 4
        else:
            return 3
    elif normalized_access in yes_answers:
        if normalized_armed_disruption in yes_answers:
            return 5
        elif normalized_idp_disruption in yes_answers:
            return 4
        elif normalized_teacher_disruption in yes_answers:
            return 3
        else:
            return 2
    return None  # Default fallback in case none of the conditions are met


##--------------------------------------------------------------------------------------------
def assign_dimension_pin(access, severity):
    # Normalize access status
    def normalize(input_string):
        if isinstance(input_string, str):
            return input_string.lower()
        return ""  # Default to empty string if input is not a string
    
    # Normalize the input to handle different cases and languages
    normalized_access = normalize(access)

    # Normalize to handle English and French variations of "yes" and "no"
    yes_answers = ['yes', 'oui']
    no_answers = ['no', 'non']

    # Mapping severity to dimension labels
    if normalized_access in no_answers:
        if severity in [4, 5]: return 'aggravating circumstances'
        elif severity == 3: return 'access'
    elif normalized_access in yes_answers:
        if severity == 3: return 'learning condition'
        if severity in [4, 5]: return 'protected environment'    
        if severity == 2: return 'not falling within the PiN dimensions'   
    
    return None  # Default fallback in case none of the conditions are met         

##--------------------------------------------------------------------------------------------
def print_subtables(severity_admin_status, pop_group_var):
    # Get the level number for pop_group_var
    level_number = severity_admin_status.index.names.index(pop_group_var)
    
    # Get unique groups
    unique_groups = severity_admin_status.index.get_level_values(level_number).unique()
    
    # Iterate and print subtables
    for group in unique_groups:
        subtable = severity_admin_status.xs(group, level=level_number)
        print(f"\nSubtable for {pop_group_var} = {group}")
        print(subtable)
        print("\n" + "-"*50 + "\n")

##--------------------------------------------------------------------------------------------
def save_subtables_to_excel(severity_admin_status, pop_group_var, file_path):
    # Get the level number for pop_group_var
    level_number = severity_admin_status.index.names.index(pop_group_var)
    
    # Get unique groups
    unique_groups = severity_admin_status.index.get_level_values(level_number).unique()
    
    # Create a Pandas Excel writer using XlsxWriter as the engine
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        # Iterate and save subtables
        for group in unique_groups:
            subtable = severity_admin_status.xs(group, level=level_number)
            subtable.to_excel(writer, sheet_name=f"{pop_group_var}_{group}")
            print(f"-------------------- Subtable for {pop_group_var} = {group} saved to sheet {pop_group_var}_{group}")
##--------------------------------------------------------------------------------------------        
def map_template_to_status(template_values, suggestions_mapping, status_values):
    results = {}
    for template in template_values:
        suggestions = suggestions_mapping.get(template, [])
        # Search for the first matching status with direct comparisons
        match = next((status for status in status_values if status in suggestions), None)
        if match:
            results[template] = match
        else:
            results[template] = 'No match found'
    return results

##--------------------------------------------------------------------------------------------        
def extract_status_data(ocha_data, mapped_statuses, pop_group_var):
    # Data frames dictionary to store each category's DataFrame
    data_frames = {}
    
    for category, status in mapped_statuses.items():
        if status != 'No match found':
            # Use the status as the category name for clarity and direct mapping
            category_name = status  # This changes the category name to the matched status value
            
            # Prepare the column names to extract based on the matched status
            children_col = f"{category} -- Children/Enfants (5-17)"
            
            # Check if these columns exist in the DataFrame
            if all(col in ocha_data.columns for col in [children_col]):
                # Create a new DataFrame for this category using the status as the category name
                category_df = ocha_data[['Admin', 'Admin Pcode', children_col]].copy()
                category_df.rename(columns={
                    children_col: 'TotN'
                }, inplace=True)
                category_df['Category'] = category_name  # Set the category name to the matched status
                category_df[pop_group_var] = status
                data_frames[category_name] = category_df
            else:
                print(f"Columns for {category} not found in OCHA data.")
        else:
            print(f"No match found for the category: {category}, skipping data extraction for this category.")

    return data_frames
##--------------------------------------------------------------------------------------------
def extract_subtables(df, pop_group_var):
    # Ensure the DataFrame has a MultiIndex and includes the population group variable
    if not isinstance(df.index, pd.MultiIndex) or pop_group_var not in df.index.names:
        raise ValueError("DataFrame must have a MultiIndex and include the specified population group variable.")

    # Get unique population groups from the specified level of the index
    unique_groups = df.index.get_level_values(pop_group_var).unique()

    # Dictionary to store each sub-DataFrame
    subtables_dict = {}

    # Extract subtables for each unique group
    for group in unique_groups:
        # Extract data for the current group
        sub_df = df.xs(group, level=pop_group_var)
        
        # Reset the index to turn MultiIndex into regular columns
        sub_df = sub_df.reset_index()
        
        # Set new DataFrame with simplified headers
        subtables_dict[group] = sub_df.rename(columns=lambda x: x if isinstance(x, str) else str(x))

    return subtables_dict

##--------------------------------------------------------------------------------------------
def calculate_category_factors(df, total_col, category_col, category_name):
    """
    Calculate ratios for specific categories and filter the DataFrame to include only necessary columns.
    Returns a dictionary of DataFrames.
    """
    result_df = df.copy()
    result_df[category_name] = result_df[category_col] / result_df[total_col].replace(0, pd.NA)
    result_df['Category'] = category_name
    columns_to_keep = ['Admin', 'Admin Pcode', category_name, 'Category']
    
    # Wrap the result in a dictionary using category_name as the key
    return {category_name: result_df[columns_to_keep]}

##--------------------------------------------------------------------------------------------
def calculate_cycle_factors(df, factor_cycle, primary_start, secondary_end, vector_cycle, single_cycle):

    if single_cycle:
        factor_cycle[0] = (vector_cycle[0] - primary_start + 1) / (secondary_end - primary_start + 1)
        factor_cycle[1] =0
        factor_cycle[2] =  (secondary_end - vector_cycle[0]) / (secondary_end - primary_start + 1)
    else:
        factor_cycle[0] = (vector_cycle[0] - primary_start + 1) / (secondary_end - primary_start + 1)
        factor_cycle[1] = (vector_cycle[1] - vector_cycle[0]) / (secondary_end - primary_start + 1)
        factor_cycle[2] = (secondary_end - vector_cycle[1]) / (secondary_end - primary_start + 1)

    # Create dictionaries to hold the categories and their respective factors
    categories = {
        'primary': factor_cycle[0],
        'upper_primary': factor_cycle[1],
        'secondary': factor_cycle[2]
    }
    # Create DataFrames for each category
    result = {}
    for category, factor in categories.items():
        temp_df = df.copy()
        temp_df[category] = factor
        temp_df['Category'] = category
        columns_to_keep = ['Admin', 'Admin Pcode', category, 'Category']
        result[category] = temp_df[columns_to_keep]
    return result
##--------------------------------------------------------------------------------------------
def reduce_index(df, level):
    df.columns = df.columns.get_level_values(1)
    df=df.droplevel(0, axis=0) 
    df=df.droplevel(0, axis=0) 
    if level == 0: df = df.reset_index( level = [0 , 1] ) 
    if level == 1: df = df.reset_index( level = [0 , 1, 2] ) 

    # Splitting the DataFrame based on pop_group_var
    groups = df.groupby(pop_group_var)
    df_list = {name: group for name, group in groups}

    return df_list

##--------------------------------------------------------------------------------------------
def add_disability_factor(df,factor=0.1, category = 'Disability'):
    # Copy the dataframe to avoid altering the original data
    result_df = df.copy()
    
    # Create the disability factor column
    category_name = category
    result_df[category_name] = factor
    
    # Add a category label
    result_df['Category'] = category_name
    
    # Define columns to keep
    columns_to_keep = ['Admin', 'Admin Pcode', category_name, 'Category']
    
    # Return the filtered DataFrame
    return {category_name: result_df[columns_to_keep]}



##--------------------------------------------------------------------------------------------
# %PiN AND #PiN PER ADMIN AND POPULATION GROUP for the strata: GENDER, SCHOOL-CYCLE 
def adjust_pin_by_strata_factor(pin_df, factor_df, category_label, tot_column):
    # Merge the pin DataFrame with the factor DataFrame on the 'Admin_2' column
    factorized_df = pd.merge(pin_df, factor_df, left_on=admin_var, right_on="Admin", how='left')
    factorized_df = pd.merge(
        pin_df, factor_df, 
        left_on=[admin_var, 'Admin Pcode'], 
        right_on=["Admin", 'Admin Pcode'], 
        how='left'
    )
   # Columns that need to be adjusted by the factor
    columns_to_adjust = [col for col in factorized_df.columns if col.startswith('#') or col == tot_column]
    del factorized_df['Admin']

    # Apply the multiplication for each column that needs adjustment
    for col in columns_to_adjust:
        factorized_df[col] *= factorized_df[category_label]

    # Drop the now unneeded factor column
    factorized_df.drop(columns=[category_label], inplace=True)
    return factorized_df



##--------------------------------------------------------------------------------------------
# preparation for overview--> SUM all the admin per population group and per strata 
def collapse_and_summarize(pin_per_admin_status_strata, category_str):
    collapsed_results = {}
    
    # Iterate over the input dictionary
    for category, df in pin_per_admin_status_strata.items():
        # Create a copy of the first row to preserve the structure
        summed_df = df.iloc[0:1].copy()

        # Identify columns to skip from summation and columns to set to zero
        columns_to_skip = [col for col in df.columns if col.startswith('%') or col == admin_var or col == 'Admin Pcode' or col == 'Population group' or col == 'Category' or col == 'Area severity']
        columns_to_zero = [col for col in df.columns if col.startswith('%')]

        # Sum all numerical columns except the skipped ones
        for col in df.columns:
            if col not in columns_to_skip:
                summed_df[col] = df[col].sum()

        # Set non-sum columns with fixed values
        summed_df['admin_2'] = 'whole country'
        summed_df['Admin Pcode'] = '0'
        summed_df['Population group'] = category
        if 'Area severity' in summed_df.columns:
            del summed_df['Area severity']

        # Set percentage columns to zero
        for col in columns_to_zero:
            summed_df[col] = 0

        # Add the modified DataFrame to the results dictionary
        summed_df = summed_df.iloc[:1]
        collapsed_results[category] = summed_df


    # Initialize the overview DataFrame with the first entry
    first_key = next(iter(collapsed_results))  # Get the first key from the dictionary
    overview_strata = collapsed_results[first_key].copy()

    # Iterate through all DataFrames in the dictionary and add their values to the overview DataFrame
    for category, df in collapsed_results.items():
        if category != first_key:  # Skip the initial DataFrame used for initialization
            overview_strata += df

    # Set final summary values
    overview_strata[admin_var] = 'Whole country'
    overview_strata['Admin Pcode'] = 0
    overview_strata['Population group'] = 'All population groups'
    overview_strata['Category'] = category_str

    overview_strata[label_perc_tot] = 0
    overview_strata[label_tot] = (overview_strata[label_tot3] +
                               overview_strata[label_tot4] +
                               overview_strata[label_tot5])

    cols = list(overview_strata.columns)
    cols.insert(cols.index(label_tot) + 1, cols.pop(cols.index('Category')))
    overview_strata = overview_strata[cols]

    return overview_strata


##--------------------------------------------------------------------------------------------
# preparation for overview--> SUM all the admin per population group and per strata 
def collapse_and_summarize_dimension(pin_per_admin_status_strata, category_str):
    collapsed_results = {}
    
    # Iterate over the input dictionary
    for category, df in pin_per_admin_status_strata.items():
        # Create a copy of the first row to preserve the structure
        summed_df = df.iloc[0:1].copy()

        # Identify columns to skip from summation and columns to set to zero
        columns_to_skip = [col for col in df.columns if col.startswith('%') or col == admin_var or col == 'Admin Pcode' or col == 'Population group' or col == 'Category' or col == 'Area severity']
        columns_to_zero = [col for col in df.columns if col.startswith('%')]

        # Sum all numerical columns except the skipped ones
        for col in df.columns:
            if col not in columns_to_skip:
                summed_df[col] = df[col].sum()

        # Set non-sum columns with fixed values
        summed_df['admin_2'] = 'whole country'
        summed_df['Admin Pcode'] = '0'
        summed_df['Population group'] = category
        if 'Area severity' in summed_df.columns:
            del summed_df['Area severity']

        # Set percentage columns to zero
        for col in columns_to_zero:
            summed_df[col] = 0

        # Add the modified DataFrame to the results dictionary
        summed_df = summed_df.iloc[:1]
        collapsed_results[category] = summed_df


    # Initialize the overview DataFrame with the first entry
    first_key = next(iter(collapsed_results))  # Get the first key from the dictionary
    overview_strata = collapsed_results[first_key].copy()

    # Iterate through all DataFrames in the dictionary and add their values to the overview DataFrame
    for category, df in collapsed_results.items():
        if category != first_key:  # Skip the initial DataFrame used for initialization
            overview_strata += df

    # Set final summary values
    overview_strata[admin_var] = 'Whole country'
    overview_strata['Admin Pcode'] = 0
    overview_strata['Population group'] = 'All population groups'
    overview_strata['Category'] = category_str

    overview_strata[label_dimension_perc_tot] = 0
    overview_strata[label_dimension_tot] = (overview_strata[label_tot_acc] +
                               overview_strata[label_tot_agg] +
                               overview_strata[label_tot_penv] + overview_strata[label_tot_lc])

    cols = list(overview_strata.columns)
    cols.insert(cols.index(label_dimension_tot) + 1, cols.pop(cols.index('Category')))
    overview_strata = overview_strata[cols]

    return overview_strata



##--------------------------------------------------------------------------------------------
def apply_formatting(file_path, color_mapping, alignment_columns):
    # Load the workbook and iterate through sheets
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        # Add empty rows at the top
        ws.insert_rows(1, 4)
        # Add empty columns to the left
        ws.insert_cols(1, 4)

        # Add the sheet name as a title in the first row
        title = ws.title
        if ws.title != "PiN TOTAL":
            title += " (5-17 y.o.)"
        max_col = ws.max_column
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=5)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if cell.value and not isinstance(cell, MergedCell):
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                        column_letter = cell.column_letter  # Get the column letter
                    except Exception as e:
                        print(f"Error: {e}, Cell: {cell}")

            if column_letter:
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width

        # Bold specific columns
        for row in ws.iter_rows(min_row=6, max_col=ws.max_column, max_row=ws.max_row):  # Start from the data row
            for cell in row:
                col_name = ws.cell(row=5, column=cell.column).value  # Row 5 contains the headers
                if col_name in [label_perc_tot, label_admin_severity, label_tot, admin_var]:
                    cell.font = Font(bold=True)  # Apply bold font

        # Check if the worksheet is "PiN TOTAL"
        if ws.title == "PiN TOTAL":
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value
                    if col_name in alignment_columns:
                        cell.alignment = Alignment(horizontal='right', vertical='center')

            # Iterate through the rows starting from the first data row
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                strata_value = row[4].value  # 'Strata' column should be the 5th column after inserting 4 empty columns

                # Determine the fill color based on 'Strata' value
                if strata_value == "TOTAL (5-17 y.o.)":
                    fill_color = colors["bluepin"]
                    for cell in row:
                        cell.font = Font(color=colors["white"], bold=True)  # Set text color to white and bold
                elif strata_value in ["Girls", "Boys"]:
                    fill_color = colors["gray"]
                elif strata_value == "ECE (5 y.o.)":
                    fill_color = colors["light_pink"]
                elif "school" in strata_value.lower():
                    fill_color = colors["light_yellow"]
                elif strata_value == "Strata":
                    fill_color = colors["white"]   
                elif "disability" in strata_value.lower():
                    fill_color = colors["white"]         
                else:
                    fill_color = colors["light_blue"] 

                # Apply fill color to the entire row if a color is determined
                if fill_color:
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Set first four columns to white
                for cell in row[:4]:
                    cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")
                    cell.border = None  # No border for the first four columns

        else:
            # Apply formatting to the headers
            for cell in ws[5]:  # Header row is now the 5th row
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply color to specific columns and make borders visible
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value
                    # Apply color based on mapping
                    if col_name in color_mapping:
                        cell.fill = PatternFill(start_color=color_mapping[col_name], end_color=color_mapping[col_name], fill_type="solid", patternType='solid')
                    # Apply alignment
                    if col_name in alignment_columns:
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                    # Apply border, but skip the first four columns
                    if cell.column > 4:
                        if row[0].row == 5:  # Bold top border for header
                            cell.border = Border(
                                top=Side(style="medium"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[0]:  # Bold left border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="medium"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[-1]:  # Bold right border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="medium"),
                                bottom=Side(style="thin"),
                            )
                        elif row[0].row == ws.max_row:  # Bold bottom border for last row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="medium"),
                            )
                        else:
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                    else:
                        cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")  # White background for the first four columns
                        cell.border = None  # No border for the first four columns

        # Apply borders around the table for "PiN TOTAL"
        if ws.title == "PiN TOTAL":
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    if cell.column > 4:
                        if row[0].row == 5:  # Bold top border for header
                            cell.border = Border(
                                top=Side(style="medium"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[0]:  # Bold left border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="medium"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[-1]:  # Bold right border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="medium"),
                                bottom=Side(style="thin"),
                            )
                        elif row[0].row == ws.max_row:  # Bold bottom border for last row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="medium"),
                            )
                        else:
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )

    # Save the formatted workbook
    wb.save(file_path)



##--------------------------------------------------------------------------------------------

def apply_formatting_dimension(file_path, color_mapping, alignment_columns):
    # Load the workbook and iterate through sheets
    wb = load_workbook(file_path)
    for ws in wb.worksheets:
        # Add empty rows at the top
        ws.insert_rows(1, 4)
        # Add empty columns to the left
        ws.insert_cols(1, 4)

        # Add the sheet name as a title in the first row
        title = ws.title
        if ws.title != "By dimension TOTAL":
            title += " (5-17 y.o.)"
        max_col = ws.max_column
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=5)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths based on content
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if cell.value and not isinstance(cell, MergedCell):
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                        column_letter = cell.column_letter  # Get the column letter
                    except Exception as e:
                        print(f"Error: {e}, Cell: {cell}")

            if column_letter:
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width

        # Bold specific columns
        for row in ws.iter_rows(min_row=6, max_col=ws.max_column, max_row=ws.max_row):  # Start from the data row
            for cell in row:
                col_name = ws.cell(row=5, column=cell.column).value  # Row 5 contains the headers
                if col_name in [label_dimension_perc_tot,  label_dimension_tot, admin_var]:
                    cell.font = Font(bold=True)  # Apply bold font

        # Check if the worksheet is "PiN TOTAL"
        if ws.title == "By dimension TOTAL":
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value
                    if col_name in alignment_columns:
                        cell.alignment = Alignment(horizontal='right', vertical='center')

            # Iterate through the rows starting from the first data row
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                strata_value = row[4].value  # 'Strata' column should be the 5th column after inserting 4 empty columns

                # Determine the fill color based on 'Strata' value
                if strata_value == "TOTAL (5-17 y.o.)":
                    fill_color = colors_dimension["bluepin"]
                    for cell in row:
                        cell.font = Font(color=colors_dimension["white"], bold=True)  # Set text color to white and bold
                elif strata_value in ["Girls", "Boys"]:
                    fill_color = colors_dimension["gray"]
                elif strata_value == "ECE (5 y.o.)":
                    fill_color = colors_dimension["light_pink"]
                elif "school" in strata_value.lower():
                    fill_color = colors_dimension["light_yellow"]
                elif strata_value == "Strata":
                    fill_color = colors_dimension["white"]   
                elif "disability" in strata_value.lower():
                    fill_color = colors_dimension["white"]         
                else:
                    fill_color = colors_dimension["light_blue"] 

                # Apply fill color to the entire row if a color is determined
                if fill_color:
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Set first four columns to white
                for cell in row[:4]:
                    cell.fill = PatternFill(start_color=colors_dimension["white"], end_color=colors_dimension["white"], fill_type="solid")
                    cell.border = None  # No border for the first four columns

        else:
            # Apply formatting to the headers
            for cell in ws[5]:  # Header row is now the 5th row
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply color to specific columns and make borders visible
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value
                    # Apply color based on mapping
                    if col_name in color_mapping:
                        cell.fill = PatternFill(start_color=color_mapping_dimension[col_name], end_color=color_mapping_dimension[col_name], fill_type="solid", patternType='solid')
                    # Apply alignment
                    if col_name in alignment_columns:
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                    # Apply border, but skip the first four columns
                    if cell.column > 4:
                        if row[0].row == 5:  # Bold top border for header
                            cell.border = Border(
                                top=Side(style="medium"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[0]:  # Bold left border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="medium"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[-1]:  # Bold right border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="medium"),
                                bottom=Side(style="thin"),
                            )
                        elif row[0].row == ws.max_row:  # Bold bottom border for last row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="medium"),
                            )
                        else:
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                    else:
                        cell.fill = PatternFill(start_color=colors_dimension["white"], end_color=colors_dimension["white"], fill_type="solid")  # White background for the first four columns
                        cell.border = None  # No border for the first four columns

        # Apply borders around the table for "PiN TOTAL"
        if ws.title == "By dimension TOTAL":
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    if cell.column > 4:
                        if row[0].row == 5:  # Bold top border for header
                            cell.border = Border(
                                top=Side(style="medium"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[0]:  # Bold left border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="medium"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[-1]:  # Bold right border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="medium"),
                                bottom=Side(style="thin"),
                            )
                        elif row[0].row == ws.max_row:  # Bold bottom border for last row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="medium"),
                            )
                        else:
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )

    # Save the formatted workbook
    wb.save(file_path)








##--------------------------------------------------------------------------------------------
# what should arrive from the user selection
admin_target = 'Admin_2: Regions'
pop_group_var = 'place_of_origin'
access_var = 'edu_access'
teacher_disruption_var = 'edu_disrupted_teacher'
idp_disruption_var = 'edu_disrupted_displaced'
armed_disruption_var = 'edu_disrupted_occupation'
barrier_var = 'edu_barrier'
selected_severity_4_barriers = [
    "Protection risks whilst at the school ",
    "Protection risks whilst travelling to the school ",
    "Child needs to work at home or on the household's own farm (i.e. is not earning an income for these activities, but may allow other family members to earn an income) ",
    "Child participating in income generating activities outside of the home",
    "Marriage, engagement and/or pregnancy",
    "Unable to enroll in school due to lack of documentation",
    "Discrimination or stigmatization of the child for any reason"]
selected_severity_5_barriers = ["Child is associated with armed forces or armed groups "]
age_var = 'edu_ind_age'
gender_var = 'edu_ind_gender'
start_month = 'september'
country= 'Somalia -- SOM'


vector_cycle = [9,14]
single_cycle = (vector_cycle[1] == 0)
primary_start = 6
secondary_end = 17

##--------------------------------------------------------------------------------------------
## status definition/suggestion:
host_suggestion = ["always_lived",'Host Community','host_communi', "always_lived","non_displaced_vulnerable",'host',"non_pdi","hote","menage_n_deplace","menage_n_deplace","resident","lebanese","Populationnondéplacée","ocap","non_deplacee","Residents","yes","4"]
IDP_suggestion = ["displaced", 'New IDPs','pdi', 'idp', 'site', 'camp', 'migrant', 'Out-of-camp', 'In-camp','no', 'pdi_site', 'pdi_fam', '2', '1' ]
returnee_suggestion = ['displaced_previously' ,'cb_returnee','ret','Returnee HH','returnee' ,'ukrainian moldovan','Returnees','5']
refugee_suggestion = ['refugees', 'refugee', 'prl', 'refugiee', '3']
ndsp_suggestion = ['ndsp','Protracted IDPs']
status_to_be_excluded = ['dnk', 'other', 'pnta', 'dont_know', 'no_answer', 'prefer_not_to_answer', 'pnpr', 'nsp', 'autre', 'do_not_know', 'decline']
template_values = ['Host/Hôte',	'IDP/PDI',	'Returnees/Retournés', 'Refugees/Refugiee', 'Other']
suggestions_mapping = {
    'Host/Hôte': host_suggestion,
    'IDP/PDI': IDP_suggestion,
    'Returnees/Retournés': returnee_suggestion,
    'Refugees/Refugiee': refugee_suggestion,
    'Other': ndsp_suggestion
}
##--------------------------------------------------------------------------------------------
##--------------------------------------------------------------------------------------------
##--------------------------------------------------------------------------------------------

# Path to your Excel file
excel_path = 'input/REACH_MSNA_2024_clean dataset_template_final.xlsx'
excel_path_ocha = 'input/ocha_pop.xlsx'

# Load the Excel file
xls = pd.ExcelFile(excel_path, engine='openpyxl')
# Print all sheet names (optional)
print(xls.sheet_names)
# Dictionary to hold your dataframes
dfs = {}
# Read each sheet into a dataframe
for sheet_name in xls.sheet_names:
    dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)

# Access specific dataframes
edu_data = dfs['edu_ind data']
household_data = dfs['SOM2404_MSNA_Tool Data ']
survey = dfs['survey']
choices = dfs['choices']

ocha_pop_data = pd.read_excel(pd.ExcelFile(excel_path_ocha, engine='openpyxl') )




####### ** 1 **       ------------------------------ manipulation and join between H and edu data   ------------------------------------------     #######
household_data['weight'] = 1
# Find the UUID columns, assuming they exist and taking only the first match for simplicity
edu_uuid_column = [col for col in edu_data.columns if 'uuid' in col.lower()][0]  # Take the first item directly
household_uuid_column = [col for col in household_data.columns if 'uuid' in col.lower()][0]  # Take the first item directly


# Extract the month from the 'start_time' column
household_data['start'] = pd.to_datetime(household_data['start'])
household_data['month'] = household_data['start'].dt.month

# Find the most similar column to "Admin2" in household_data
admin_var = process.extractOne(admin_target, household_data.columns.tolist())[0]  # Take the string directly

# Columns to include in the merge
columns_to_include = [household_uuid_column, admin_var, pop_group_var, 'month', 'weights', 'weight']
edu_data = edu_data.drop(columns=[col for col in columns_to_include if col in edu_data.columns], errors='ignore')

# ----> Perform the joint_by
edu_data = pd.merge(edu_data, household_data[columns_to_include], left_on=edu_uuid_column, right_on=household_uuid_column, how='left')

##refining for school age-children
#edu_data = edu_data[(edu_data[age_var] >= 5) & (edu_data[age_var] <= 18)]

edu_data['edu_age_corrected'] = edu_data.apply(lambda row: row[age_var] - 1 if calculate_age_correction(start_month, row['month']) else row[age_var], axis=1)
edu_data['school_cycle'] = edu_data['edu_age_corrected'].apply(
    lambda x: assign_school_cycle(
        x, 
        single_cycle=single_cycle, 
        lower_primary_start_var=primary_start, 
        lower_primary_end_var=vector_cycle[0], 
        upper_primary_end_var=vector_cycle[1] if not single_cycle else None
    )
)
edu_data = edu_data[(edu_data['edu_age_corrected'] >= 5) & (edu_data['edu_age_corrected'] <= 17)]


####### ** 2 **       ------------------------------ severity definition and calculation ------------------------------------------     #######
severity_4_matches = find_matching_choices(choices, selected_severity_4_barriers)
severity_5_matches = find_matching_choices(choices, selected_severity_5_barriers)
names_severity_4 = [entry['name'] for entry in severity_4_matches]
names_severity_5 = [entry['name'] for entry in severity_5_matches]

# Apply the conditions and choices to create the new 'severity_category' column
edu_data['severity_category'] = edu_data.apply(lambda row: calculate_severity(
    access=row[access_var], 
    barrier=row[barrier_var], 
    armed_disruption=row[armed_disruption_var], 
    idp_disruption=row[idp_disruption_var], 
    teacher_disruption=row[teacher_disruption_var], 
    names_severity_4=names_severity_4, 
    names_severity_5=names_severity_5
), axis=1)

# Add the new column 'dimension_pin' to edu_data
edu_data['dimension_pin'] = edu_data.apply(lambda row: assign_dimension_pin(
    access=row[access_var],
    severity= row['severity_category']
    ), axis=1)

####### ** 3 **       ------------------------------ Analysis per ADMIN AND POPULATION GROUP ------------------------------------------     #######
df = pd.DataFrame(edu_data)
startum_gender = edu_data[gender_var]
startum_school_cycle = edu_data['school_cycle']
print('               -------- GENDER DISAGGREGATION  ---------           ')
severity_by_gender = df.groupby([admin_var, pop_group_var,gender_var, 'severity_category']).agg(
    total_weight=('weights', 'sum')
).groupby(level=[0, 1]).apply(
    lambda x: x / x.sum()
).unstack(fill_value=0)
print(severity_by_gender)

print('                             ')
print('               -------- school-cycle DISAGGREGATION  ---------           ')
severity_by_cycle = df.groupby([admin_var, pop_group_var,startum_school_cycle, 'severity_category']).agg(
    total_weight=('weights', 'sum')
).groupby(level=[0, 1]).apply(
    lambda x: x / x.sum()
).unstack(fill_value=0)
print(severity_by_cycle)

print('                             ')
print('            -------    CORRECT PIN    -------             ')
severity_admin_status = df.groupby([admin_var, pop_group_var, 'severity_category']).agg(
    total_weight=('weights', 'sum')
).groupby(level=[0, 1]).apply(
    lambda x: x / x.sum()
).unstack(fill_value=0)
print(severity_admin_status)
print('                             ')
print('            -------    CORRECT TARGETTING    -------             ')
dimension_admin_status = df.groupby([admin_var, pop_group_var, 'dimension_pin']).agg(
    total_weight=('weights', 'sum')
).groupby(level=[0, 1]).apply(
    lambda x: x / x.sum()
).unstack(fill_value=0)
print(dimension_admin_status)


## reducing the multiindex of the panda dataframe
severity_admin_status_list = reduce_index(severity_admin_status, 0)
severity_by_gender_list = reduce_index(severity_by_gender, 1)
severity_by_cycle_list = reduce_index(severity_by_cycle, 1)

dimension_admin_status_list = reduce_index(dimension_admin_status, 0)

####### ** 4 **       ------------------------------ matching between the admin and the ocha population data ------------------------------------------     #######

## finding the match between the OCHA status cathegory and the country status. 
status_allvalues = edu_data[pop_group_var].unique()
status_values = [status for status in edu_data[pop_group_var].unique() if status not in status_to_be_excluded]# Retrieve unique values directly without converting to lowercase
for key, suggestions in suggestions_mapping.items():
    suggestions_mapping[key] = suggestions  # keeping original case

mapped_statuses = map_template_to_status(template_values, suggestions_mapping, status_values)
category_data_frames = extract_status_data(ocha_pop_data, mapped_statuses, pop_group_var)# Extract population figures based on mapped statuses without modifying the case

# Debugging and data inspection
for key, value in mapped_statuses.items():
    print(f"{key}: {value}")
for category, df in category_data_frames.items():
    df.rename(columns={'Admin': admin_var}, inplace=True)
    print(f"Category: {category}")
    print(df.head())  # Display the first few rows of the DataFrame
    print("\n" + "-"*50 + "\n")  # Print a separator for better readability between categories


####### ** 5 **       ------------------------------ creating tables with factors for the gender and school-cycle categories ------------------------------------------     #######

## calculate the difference population group per school-cycle according to the country and the tot-children population 
factor_cycle = [0.5,0.5,0]
factor_disability = 0.1
## create table per strata
category_tot = 'All'
category_girl = 'Girl'
category_boy = 'Boy'
category_ece= 'ECE'
category_primary= 'primary'
category_upper_primary= 'upper_primary'
category_secondary= 'secondary'
category_disability = 'Disability'

children_tot_col = 'ToT -- Children/Enfants (5-17)'
girls_tot_col = 'ToT -- Girls/Filles (5-17)'
boys_tot_col = 'ToT -- Boys/Garcons (5-17)'
ece_tot_col = '5yo -- Children/Enfants'


# Calculate category factors
category_factors = {
    **calculate_category_factors(ocha_pop_data, children_tot_col, girls_tot_col, category_girl),
    **calculate_category_factors(ocha_pop_data, children_tot_col, boys_tot_col, category_boy),
    **calculate_category_factors(ocha_pop_data, children_tot_col, ece_tot_col, category_ece)
}
# Calculate factors for each school cycle
school_cycle_factors = calculate_cycle_factors(ocha_pop_data, factor_cycle, primary_start, secondary_end, vector_cycle, single_cycle)
# Combine all factors into one dictionary
factor_category = {**category_factors, **school_cycle_factors}
disability_factors = add_disability_factor(ocha_pop_data, factor_disability,category_disability )
# Update the factor_category dictionary with the new disability category
factor_category.update(disability_factors)

# Print results
for key, df in factor_category.items():
    print(f"{key} factors:\n{df}\n")


####### ** 6.A **       ------------------------------ %PiN AND #PiN PER ADMIN AND POPULATION GROUP using ocha figures ------------------------------------------     #######
int_2 = '2.0'
int_3 = '3.0'
int_4 = '4.0'
int_5 = '5.0'
label_perc2 = '% 1-2'
label_perc3 = '% 3'
label_perc4 = '% 4'
label_perc5 = '% 5'
label_tot2 = '# 1-2'
label_tot3 = '# 3'
label_tot4 = '# 4'
label_tot5 = '# 5'
label_perc_tot = '% Tot PiN (3+)'
label_tot = '# Tot PiN (3+)'
label_admin_severity = 'Area severity'
label_tot_population = 'TotN'

pin_per_admin_status = {}

# Assume category_data_frames is a dictionary of DataFrames, indexed by category
for category, df in category_data_frames.items():
    # Ensure both DataFrames are ready to merge
    if category in severity_admin_status_list:
        # Fetch the corresponding DataFrame from the grouped data
        grouped_df = severity_admin_status_list[category]     
        # Merge on specified columns
        pop_group_df = pd.merge(grouped_df, df, on=[admin_var, pop_group_var])
        pop_group_df.columns = [str(col) for col in pop_group_df.columns]

        ## arranging columns 
        cols = list(pop_group_df.columns)
        cols.remove('Admin Pcode')
        cols.insert( cols.index(admin_var) + 1, 'Admin Pcode')
        pop_group_df = pop_group_df[cols]


        ## !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!   calculation of the tot Pin and admin severity -->
        # Step 1: Create the new column with initial zeros
        pop_group_df = pop_group_df.rename(columns={
            pop_group_var: 'Population group',
            int_2: label_perc2,
            int_3: label_perc3,
            int_4: label_perc4,
            int_5: label_perc5
        })
        del pop_group_df['Category']

        # Initialize total columns with zeros
        for label in [label_tot2, label_tot3, label_tot4, label_tot5]:
            pop_group_df[label] = 0

        
        cols = list(pop_group_df.columns)
        # Move the newly added column to the desired position
        cols.insert(cols.index(label_perc2) + 1, cols.pop(cols.index(label_tot2)))
        cols.insert(cols.index(label_perc3) + 1, cols.pop(cols.index(label_tot3)))
        cols.insert(cols.index(label_perc4) + 1, cols.pop(cols.index(label_tot4)))
        cols.insert(cols.index(label_perc5) + 1, cols.pop(cols.index(label_tot5)))
        pop_group_df = pop_group_df[cols]     

        # Calculate total PiN for each severity level
        for perc_label, total_label in [(label_perc2, label_tot2), 
                                        (label_perc3, label_tot3), 
                                        (label_perc4, label_tot4), 
                                        (label_perc5, label_tot5)]:
            pop_group_df[total_label] = pop_group_df[perc_label] * pop_group_df[label_tot_population]

        
        # Reorder columns as needed
        cols = list(pop_group_df.columns)
        cols.insert(cols.index('Population group') + 1, cols.pop(cols.index(label_tot_population)))
        pop_group_df = pop_group_df[cols]     


        cols.remove(label_tot_population)
        cols.insert( cols.index('Population group') + 1, label_tot_population)
        pop_group_df = pop_group_df[cols]

        # Save modified DataFrame back into the dictionary under the category key
        pin_per_admin_status[category] = pop_group_df


# Define the output Excel file path
output_excel_path = 'output_test/pin_per_admin_status_test0.xlsx'

# Create an Excel writer object
with pd.ExcelWriter(output_excel_path) as writer:
    # Iterate over each category and DataFrame in the dictionary
    for category, df in pin_per_admin_status.items():
        # Write the DataFrame to a sheet named after the category
        df.to_excel(writer, sheet_name=category, index=False)

print(f"Data has been saved to {output_excel_path}")





####### ** 6.B **       ------------------------------ %dimension AND #dimension PER ADMIN AND POPULATION GROUP using ocha figures ------------------------------------------     #######
int_acc = 'access'
int_agg= 'aggravating circumstances'
int_lc = 'learning condition'
int_penv = 'protected environment'
int_out = 'not falling within the PiN dimensions'
label_perc_acc = '% Access'
label_perc_agg= '% Aggravating circumstances'
label_perc_lc = '% Learning conditions'
label_perc_penv = '% Protected environment'
label_perc_out = '% Not falling within the PiN dimensions'
label_tot_acc = '# Access'
label_tot_agg= '# Aggravating circumstances'
label_tot_lc = '# Learning conditions'
label_tot_penv = '# Protected environment'
label_tot_out = '# Not falling within the PiN dimensions'

label_dimension_perc_tot = '% Tot in PiN Dimensions'
label_dimension_tot = '# Tot in PiN Dimensions'

label_dimension_tot_population = 'TotN'

dimension_per_admin_status = {}

# Assume category_data_frames is a dictionary of DataFrames, indexed by category
for category, df in category_data_frames.items():
    # Ensure both DataFrames are ready to merge
    if category in dimension_admin_status_list:
        # Fetch the corresponding DataFrame from the grouped data
        grouped_df = dimension_admin_status_list[category]     
        # Merge on specified columns
        pop_group_df = pd.merge(grouped_df, df, on=[admin_var, pop_group_var])
        pop_group_df.columns = [str(col) for col in pop_group_df.columns]

        ## arranging columns 
        cols = list(pop_group_df.columns)
        cols.remove('Admin Pcode')
        cols.insert( cols.index(admin_var) + 1, 'Admin Pcode')
        pop_group_df = pop_group_df[cols]


        ## !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!   calculation of the tot Pin and admin severity -->
        # Step 1: Create the new column with initial zeros
        pop_group_df = pop_group_df.rename(columns={
            pop_group_var: 'Population group',
            int_acc: label_perc_acc,
            int_agg: label_perc_agg,
            int_lc: label_perc_lc,
            int_penv: label_perc_penv,
            int_out: label_perc_out
        })
        del pop_group_df['Category']

        # Initialize total columns with zeros
        for label in [label_tot_acc, label_tot_agg, label_tot_lc, label_tot_penv, label_tot_out]:
            pop_group_df[label] = 0


        cols = list(pop_group_df.columns)
        cols.remove(label_perc_out)
        cols.insert( cols.index(label_perc_penv) + 1, label_perc_out)
        pop_group_df = pop_group_df[cols]

    
        cols = list(pop_group_df.columns)
        # Move the newly added column to the desired position
        
        cols.insert(cols.index(label_perc_acc) + 1, cols.pop(cols.index(label_tot_acc)))
        cols.insert(cols.index(label_perc_agg) + 1, cols.pop(cols.index(label_tot_agg)))
        cols.insert(cols.index(label_perc_lc) + 1, cols.pop(cols.index(label_tot_lc)))
        cols.insert(cols.index(label_perc_penv) + 1, cols.pop(cols.index(label_tot_penv)))
        cols.insert(cols.index(label_perc_out) + 1, cols.pop(cols.index(label_tot_out)))
        pop_group_df = pop_group_df[cols]


        for perc_label, total_label in [(label_perc_acc, label_tot_acc), 
                                        (label_perc_agg, label_tot_agg), 
                                        (label_perc_lc, label_tot_lc), 
                                        (label_perc_penv, label_tot_penv),
                                        (label_perc_out, label_tot_out)]:
            pop_group_df[total_label] = pop_group_df[perc_label] * pop_group_df[label_dimension_tot_population]
        # Save modified DataFrame back into the dictionary under the category key

                
        # Reorder columns as needed
        cols = list(pop_group_df.columns)
        cols.insert(cols.index('Population group') + 1, cols.pop(cols.index(label_dimension_tot_population)))
        pop_group_df = pop_group_df[cols]     


        cols.remove(label_dimension_tot_population)
        cols.insert( cols.index('Population group') + 1, label_dimension_tot_population)
        pop_group_df = pop_group_df[cols]
        dimension_per_admin_status[category] = pop_group_df



# Define the output Excel file path
output_excel_dimension_path = 'output_test/dimension_per_admin_status_test0.xlsx'

# Create an Excel writer object
with pd.ExcelWriter(output_excel_dimension_path) as writer:
    # Iterate over each category and DataFrame in the dictionary
    for category, df in dimension_per_admin_status.items():
        # Write the DataFrame to a sheet named after the category
        df.to_excel(writer, sheet_name=category, index=False)

print(f"Data has been saved to {output_excel_dimension_path}")








####### ** 7 **       ------------------------------ %PiN AND #PiN PER ADMIN AND POPULATION GROUP for the strata: GENDER, SCHOOL-CYCLE ------------------------------------------     #######
## PiN
pin_per_admin_status_girl = {}
pin_per_admin_status_boy = {}
pin_per_admin_status_ece = {}
pin_per_admin_status_primary = {}
pin_per_admin_status_upper_primary = {}
pin_per_admin_status_secondary = {}
pin_per_admin_status_disabilty = {}


for category, df in pin_per_admin_status.items():
    pin_per_admin_status_girl[category] = adjust_pin_by_strata_factor(df, factor_category[category_girl], category_girl, tot_column= label_tot_population)
    pin_per_admin_status_boy[category] = adjust_pin_by_strata_factor(df, factor_category[category_boy], category_boy, tot_column= label_tot_population)
    pin_per_admin_status_ece[category] = adjust_pin_by_strata_factor(df, factor_category[category_ece], category_ece, tot_column= label_tot_population)
    pin_per_admin_status_primary[category] = adjust_pin_by_strata_factor(df, factor_category[category_primary], category_primary, tot_column= label_tot_population)
    pin_per_admin_status_upper_primary[category] = adjust_pin_by_strata_factor(df, factor_category[category_upper_primary], category_upper_primary, tot_column= label_tot_population)
    pin_per_admin_status_secondary[category] = adjust_pin_by_strata_factor(df, factor_category[category_secondary], category_secondary, tot_column= label_tot_population)
    pin_per_admin_status_disabilty[category] = adjust_pin_by_strata_factor(df, factor_category[category_disability], category_disability, tot_column= label_tot_population)


## dimension
dimension_per_admin_status_girl = {}
dimension_per_admin_status_boy = {}
dimension_per_admin_status_ece = {}
dimension_per_admin_status_primary = {}
dimension_per_admin_status_upper_primary = {}
dimension_per_admin_status_secondary = {}
dimension_per_admin_status_disabilty = {}


for category, df in dimension_per_admin_status.items():
    dimension_per_admin_status_girl[category] = adjust_pin_by_strata_factor(df, factor_category[category_girl], category_girl, tot_column= label_dimension_tot_population)
    dimension_per_admin_status_boy[category] = adjust_pin_by_strata_factor(df, factor_category[category_boy], category_boy, tot_column= label_dimension_tot_population)
    dimension_per_admin_status_ece[category] = adjust_pin_by_strata_factor(df, factor_category[category_ece], category_ece, tot_column= label_dimension_tot_population)
    dimension_per_admin_status_primary[category] = adjust_pin_by_strata_factor(df, factor_category[category_primary], category_primary, tot_column= label_dimension_tot_population)
    dimension_per_admin_status_upper_primary[category] = adjust_pin_by_strata_factor(df, factor_category[category_upper_primary], category_upper_primary, tot_column= label_dimension_tot_population)
    dimension_per_admin_status_secondary[category] = adjust_pin_by_strata_factor(df, factor_category[category_secondary], category_secondary, tot_column= label_dimension_tot_population)
    dimension_per_admin_status_disabilty[category] = adjust_pin_by_strata_factor(df, factor_category[category_disability], category_disability, tot_column= label_dimension_tot_population)



####### ** 8.A **       ------------------------------ calculate tot PiN --> 3+ and admin severity for pin_per_admin_status ------------------------------------------     #######
Tot_PiN_JIAF = pin_per_admin_status


# Iterate over the pin_per_admin_status dictionary to apply the new operations
for category, pop_group_df in Tot_PiN_JIAF.items():
    # Initialize new columns for percentage total, total PiN, and admin severity
    pop_group_df[label_perc_tot] = 0
    pop_group_df[label_tot] = 0
    pop_group_df[label_admin_severity] = 0

    # Reorder columns to place new columns at desired positions
    cols = list(pop_group_df.columns)
    cols.insert(cols.index(label_tot5) + 1, cols.pop(cols.index(label_perc_tot)))
    cols.insert(cols.index(label_perc_tot) + 1, cols.pop(cols.index(label_tot)))
    cols.insert(cols.index(label_tot) + 1, cols.pop(cols.index(label_admin_severity)))
    pop_group_df = pop_group_df[cols]

    # Calculate the total percentage and total PiN for severity levels 3+
    pop_group_df[label_perc_tot] = (pop_group_df[label_perc3] +
                                    pop_group_df[label_perc4] +
                                    pop_group_df[label_perc5])

    pop_group_df[label_tot] = (pop_group_df[label_tot3] +
                               pop_group_df[label_tot4] +
                               pop_group_df[label_tot5])

    # Define conditions based on specified logic
    conditions = [
        pop_group_df[label_perc5] > 0.2,
        (pop_group_df[label_perc5] + pop_group_df[label_perc4]) > 0.2,
        (pop_group_df[label_perc5] + pop_group_df[label_perc4] + pop_group_df[label_perc3]) > 0.2,
        (pop_group_df[label_perc5] + pop_group_df[label_perc4] + pop_group_df[label_perc3] + pop_group_df[label_perc2]) > 0.2
    ]

    # Corresponding values for each condition
    choices = ['5', '4', '3', '1-2']

    # Apply the conditions to determine admin severity
    pop_group_df[label_admin_severity] = np.select(conditions, choices, default='0')

    # Save the updated DataFrame back to the dictionary
    Tot_PiN_JIAF[category] = pop_group_df


# Define the output Excel file path
output_excel_path_JIAF = 'output_test/pin_per_admin_status_test_JIAF.xlsx'

# Create an Excel writer object
with pd.ExcelWriter(output_excel_path_JIAF) as writer:
    # Iterate over each category and DataFrame in the dictionary
    for category, df in Tot_PiN_JIAF.items():
        # Write the DataFrame to a sheet named after the category
        df.to_excel(writer, sheet_name=category, index=False)

print(f"Data has been saved to {output_excel_path_JIAF}")


####### ** 8.B **       ------------------------------ calculate tot PiN --> 3+ and admin severity for pin_per_admin_status ------------------------------------------     #######

Tot_Dimension_JIAF = dimension_per_admin_status


# Iterate over the pin_per_admin_status dictionary to apply the new operations
for category, pop_group_df in Tot_Dimension_JIAF.items():
    # Initialize new columns for percentage total, total PiN, and admin severity
    pop_group_df[label_dimension_perc_tot] = 0
    pop_group_df[label_dimension_tot] = 0

    # Reorder columns to place new columns at desired positions
    cols = list(pop_group_df.columns)
    cols.insert(cols.index(label_tot_out) + 1, cols.pop(cols.index(label_dimension_perc_tot)))
    cols.insert(cols.index(label_dimension_perc_tot) + 1, cols.pop(cols.index(label_dimension_tot)))
    pop_group_df = pop_group_df[cols]

    # Calculate the total percentage and total PiN for severity levels 3+
    pop_group_df[label_dimension_perc_tot] = (pop_group_df[label_perc_acc] +
                                    pop_group_df[label_perc_agg] +
                                    pop_group_df[label_perc_lc] +
                                    pop_group_df[label_perc_penv])

    pop_group_df[label_dimension_tot] = (pop_group_df[label_tot_acc] +
                               pop_group_df[label_tot_agg] +
                               pop_group_df[label_tot_lc]+
                               pop_group_df[label_tot_penv])


    # Save the updated DataFrame back to the dictionary
    Tot_Dimension_JIAF[category] = pop_group_df


# Define the output Excel file path
output_excel_path_dimension_JIAF = 'output_test/dimension_per_admin_status_test_JIAF.xlsx'

# Create an Excel writer object
with pd.ExcelWriter(output_excel_path_dimension_JIAF) as writer:
    # Iterate over each category and DataFrame in the dictionary
    for category, df in Tot_Dimension_JIAF.items():
        # Write the DataFrame to a sheet named after the category
        df.to_excel(writer, sheet_name=category, index=False)

print(f"Data has been saved to {output_excel_path_dimension_JIAF}")




####### ** 9 **       ------------------------------  preparation for overview--> SUM all the admin per population group and per strata ------------------------------------------     #######
overview_ToT = collapse_and_summarize(pin_per_admin_status, 'TOTAL (5-17 y.o.)')
overview_girl = collapse_and_summarize(pin_per_admin_status_girl, 'Girls')
overview_boy = collapse_and_summarize(pin_per_admin_status_boy, 'Boys')
overview_ece = collapse_and_summarize(pin_per_admin_status_ece, 'ECE (5 y.o.)')
overview_primary = collapse_and_summarize(pin_per_admin_status_primary, 'Primary school')
overview_upper_primary = collapse_and_summarize(pin_per_admin_status_upper_primary, 'Upper primary school')
overview_secondary = collapse_and_summarize(pin_per_admin_status_secondary, 'Secondary school')
overview_disabilty = collapse_and_summarize(pin_per_admin_status_disabilty, 'Children with disability')

collapsed_results_pop = {}
for category, df in pin_per_admin_status.items():
        # Create a copy of the first row to preserve the structure
        summed_df = df.iloc[0:1].copy()

        # Identify columns to skip from summation and columns to set to zero
        columns_to_skip = [col for col in df.columns if col.startswith('%') or col == admin_var or col == 'Admin Pcode' or col == 'Population group' or col == 'Category' or col== 'Area severity']
        columns_to_zero = [col for col in df.columns if col.startswith('%')]

        # Sum all numerical columns except the skipped ones
        for col in df.columns:
            if col not in columns_to_skip:
                summed_df[col] = df[col].sum()

        # Set non-sum columns with fixed values
        summed_df['admin_2'] = 'whole country'
        summed_df['Admin Pcode'] = '0'
        summed_df['Population group'] = category
        del summed_df['Area severity']

        # Set percentage columns to zero
        for col in columns_to_zero:
            summed_df[col] = 0

        # Add the modified DataFrame to the results dictionary
        summed_df = summed_df.iloc[:1]
        collapsed_results_pop[category] = summed_df




overview_dimension_ToT = collapse_and_summarize_dimension(dimension_per_admin_status, 'TOTAL (5-17 y.o.)')
overview_dimension_girl = collapse_and_summarize_dimension(dimension_per_admin_status_girl, 'Girls')
overview_dimension_boy = collapse_and_summarize_dimension(dimension_per_admin_status_boy, 'Boys')
overview_dimension_ece = collapse_and_summarize_dimension(dimension_per_admin_status_ece, 'ECE (5 y.o.)')
overview_dimension_primary = collapse_and_summarize_dimension(dimension_per_admin_status_primary, 'Primary school')
overview_dimension_upper_primary = collapse_and_summarize_dimension(dimension_per_admin_status_upper_primary, 'Upper primary school')
overview_dimension_secondary = collapse_and_summarize_dimension(dimension_per_admin_status_secondary, 'Secondary school')
overview_dimension_disabilty = collapse_and_summarize_dimension(dimension_per_admin_status_disabilty, 'Children with disability')

collapsed_results_dimension_pop = {}
for category, df in dimension_per_admin_status.items():
        # Create a copy of the first row to preserve the structure
        summed_dimension_df = df.iloc[0:1].copy()

        # Identify columns to skip from summation and columns to set to zero
        columns_to_skip = [col for col in df.columns if col.startswith('%') or col == admin_var or col == 'Admin Pcode' or col == 'Population group' or col == 'Category' or col== 'Area severity']
        columns_to_zero = [col for col in df.columns if col.startswith('%')]

        # Sum all numerical columns except the skipped ones
        for col in df.columns:
            if col not in columns_to_skip:
                summed_dimension_df[col] = df[col].sum()

        # Set non-sum columns with fixed values
        summed_dimension_df['admin_2'] = 'whole country'
        summed_dimension_df['Admin Pcode'] = '0'
        summed_dimension_df['Population group'] = category
        if 'Area severity' in summed_dimension_df.columns:
            del summed_dimension_df['Area severity']

        # Set percentage columns to zero
        for col in columns_to_zero:
            summed_dimension_df[col] = 0

        # Add the modified DataFrame to the results dictionary
        summed_dimension_df = summed_dimension_df.iloc[:1]
        collapsed_results_dimension_pop[category] = summed_dimension_df


####### ** 10.A **       ------------------------------  Creating OVERVIEW file ------------------------------------------     #######

dfs = []
dfs.append(overview_ToT)

# Add the single-row entries from collapsed_results_pop
for category, df in collapsed_results_pop.items():
    single_row = df.iloc[0].copy()
    single_row['Category'] = f"{category} (5-17 y.o.)"
    # Convert the Series to a DataFrame with a single row and append it to the list
    single_row_df = single_row.to_frame().T
    dfs.append(single_row_df)

# Determine the strata_summarized_data based on the value of single_cycle
if single_cycle:
    strata_summarized_data = [
        overview_girl,
        overview_boy,
        overview_ece,
        overview_primary,
        overview_secondary,
        overview_disabilty
    ]
else:
    strata_summarized_data = [
        overview_girl,
        overview_boy,
        overview_ece,
        overview_primary,
        overview_upper_primary,
        overview_secondary,
        overview_disabilty
    ]

# Add the remaining summarized data to the list
dfs.extend(strata_summarized_data)
# Concatenate all DataFrames in the list into a single DataFrame
final_overview_df = pd.concat(dfs, ignore_index=True)

## organization and manipulation 
cols = list(final_overview_df.columns)
cols.insert(cols.index(admin_var) + 1, cols.pop(cols.index('Category')))
final_overview_df = final_overview_df[cols]
del final_overview_df[admin_var]
del final_overview_df['Admin Pcode']
final_overview_df = final_overview_df.rename(columns={'Category': 'Strata'})

final_overview_df[label_perc2] = final_overview_df[label_tot2]/final_overview_df[label_tot_population]
final_overview_df[label_perc3] = final_overview_df[label_tot3]/final_overview_df[label_tot_population]
final_overview_df[label_perc4] = final_overview_df[label_tot4]/final_overview_df[label_tot_population]
final_overview_df[label_perc5] = final_overview_df[label_tot5]/final_overview_df[label_tot_population]
final_overview_df[label_perc_tot] = final_overview_df[label_tot]/final_overview_df[label_tot_population]

columns_perc = [col for col in final_overview_df.columns if col.startswith('%')]
for col in final_overview_df.columns:
    if col in columns_perc:
        final_overview_df[col] = final_overview_df[col].apply(lambda x: f"{x * 100:.1f}%")




####### ** 10.B **       ------------------------------  Creating OVERVIEW file ------------------------------------------     #######

dfs = []
dfs.append(overview_dimension_ToT)

# Add the single-row entries from collapsed_results_pop
for category, df in collapsed_results_dimension_pop.items():
    single_dimension_row = df.iloc[0].copy()
    single_dimension_row['Category'] = f"{category} (5-17 y.o.)"
    # Convert the Series to a DataFrame with a single row and append it to the list
    single_row_dimension_df = single_dimension_row.to_frame().T
    dfs.append(single_row_dimension_df)

# Determine the strata_summarized_data based on the value of single_cycle
if single_cycle:
    strata_summarized_dimension_data = [
        overview_dimension_girl,
        overview_dimension_boy,
        overview_dimension_ece,
        overview_dimension_primary,
        overview_dimension_secondary,
        overview_dimension_disabilty
    ]
else:
    strata_summarized_dimension_data = [
        overview_dimension_girl,
        overview_dimension_boy,
        overview_dimension_ece,
        overview_dimension_primary,
        overview_dimension_upper_primary,
        overview_dimension_secondary,
        overview_dimension_disabilty
    ]

# Add the remaining summarized data to the list
dfs.extend(strata_summarized_dimension_data)
# Concatenate all DataFrames in the list into a single DataFrame
final_overview_dimension_df = pd.concat(dfs, ignore_index=True)

## organization and manipulation 
cols = list(final_overview_dimension_df.columns)
cols.insert(cols.index(admin_var) + 1, cols.pop(cols.index('Category')))
final_overview_dimension_df = final_overview_dimension_df[cols]
del final_overview_dimension_df[admin_var]
del final_overview_dimension_df['Admin Pcode']
final_overview_dimension_df = final_overview_dimension_df.rename(columns={'Category': 'Strata'})

final_overview_dimension_df[label_perc_acc] = final_overview_dimension_df[label_tot_acc]/final_overview_dimension_df[label_dimension_tot_population]
final_overview_dimension_df[label_perc_agg] = final_overview_dimension_df[label_tot_agg]/final_overview_dimension_df[label_dimension_tot_population]
final_overview_dimension_df[label_perc_lc] = final_overview_dimension_df[label_tot_lc]/final_overview_dimension_df[label_dimension_tot_population]
final_overview_dimension_df[label_perc_penv] = final_overview_dimension_df[label_tot_penv]/final_overview_dimension_df[label_dimension_tot_population]
final_overview_dimension_df[label_perc_out] = final_overview_dimension_df[label_tot_out]/final_overview_dimension_df[label_dimension_tot_population]
final_overview_dimension_df[label_dimension_perc_tot] = final_overview_dimension_df[label_dimension_tot]/final_overview_dimension_df[label_dimension_tot_population]

columns_perc_dimension = [col for col in final_overview_dimension_df.columns if col.startswith('%')]
for col in final_overview_dimension_df.columns:
    if col in columns_perc_dimension:
        final_overview_dimension_df[col] = final_overview_dimension_df[col].apply(lambda x: f"{x * 100:.1f}%")




####### ** 11 **       ------------------------------  Rounding and Saving the JIAF AND OCHA OUTPUT ------------------------------------------     #######
## ROUNDING
percentage_round = 1
figures_round = 0

for category, df in Tot_PiN_JIAF.items():
    for col in df.columns:
        if col.startswith('#'):
            df[col] = pd.to_numeric(df[col], errors='coerce').round(figures_round)  # Convert to numeric and round
        elif col.startswith('%'):
            df[col] = df[col].apply(lambda x: f"{x * 100:.1f}%")
    df[label_tot_population] = df[label_tot_population].round(figures_round)


for col in final_overview_df.columns:
    if col.startswith('#'):
        print(col)
        final_overview_df[col] = pd.to_numeric(final_overview_df[col], errors='coerce').round(figures_round)  # Convert to numeric and round
        
final_overview_df[label_tot_population] = pd.to_numeric(final_overview_df[label_tot_population], errors='coerce').round(figures_round)  # Convert to numeric and round


## dimension
for category, df in Tot_Dimension_JIAF.items():
    for col in df.columns:
        if col.startswith('#'):
            df[col] = pd.to_numeric(df[col], errors='coerce').round(figures_round)  # Convert to numeric and round
        elif col.startswith('%'):
            df[col] = df[col].apply(lambda x: f"{x * 100:.1f}%")
    df[label_dimension_tot_population] = df[label_dimension_tot_population].round(figures_round)


for col in final_overview_dimension_df.columns:
    if col.startswith('#'):
        print(col)
        final_overview_dimension_df[col] = pd.to_numeric(final_overview_dimension_df[col], errors='coerce').round(figures_round)  # Convert to numeric and round
        
final_overview_dimension_df[label_dimension_tot_population] = pd.to_numeric(final_overview_dimension_df[label_dimension_tot_population], errors='coerce').round(figures_round)  # Convert to numeric and round

## -----------------------------  saving
# Replace spaces and special characters in the country name if necessary
country_label = country.replace(" ", "_").replace("--", "_").replace("/", "_")
# Output file paths with the country included
output_JIAF = f'output/PiN_JIAF_{country_label}.xlsx'
output_OCHA = f'output/PiN_overview_OCHA_{country_label}.xlsx'


# 1.A)
with pd.ExcelWriter(output_JIAF) as writer:
    for category, df in Tot_PiN_JIAF.items():
        # Write the DataFrame to a sheet named after the category
        sheet_name = f"PIN -- {category}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"PiN output tailored for JIAF has been saved to {output_JIAF}")



# 2.A)
with pd.ExcelWriter(output_OCHA) as writer:
    # Write the final_overview_df to the first sheet named "PiN TOTAL"
    final_overview_df.to_excel(writer, sheet_name="PiN TOTAL", index=False)

    # Write each category DataFrame to a separate sheet
    for category, df in Tot_PiN_JIAF.items():
        # Write the DataFrame to a sheet named after the category
        sheet_name = f"PIN -- {category}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"PiN output tailored for OCHA has been saved to {output_OCHA}")


# Output file paths with the country included
output_dimension_JIAF = f'output/Dimension_JIAF_{country_label}.xlsx'
output_dimension_OCHA = f'output/Dimension_overview_OCHA_{country_label}.xlsx'


# 1.B)
with pd.ExcelWriter(output_dimension_JIAF) as writer:
    for category, df in Tot_Dimension_JIAF.items():
        # Write the DataFrame to a sheet named after the category
        sheet_name = f"By dimension -- {category}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Dimension PiN output tailored for JIAF has been saved to {output_dimension_JIAF}")



# 2.B)
with pd.ExcelWriter(output_dimension_OCHA) as writer:
    # Write the final_overview_df to the first sheet named "PiN TOTAL"
    final_overview_dimension_df.to_excel(writer, sheet_name="By dimension TOTAL", index=False)

    # Write each category DataFrame to a separate sheet
    for category, df in Tot_Dimension_JIAF.items():
        # Write the DataFrame to a sheet named after the category
        sheet_name = f"By dimension -- {category}"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Dimension PiN output tailored for OCHA has been saved to {output_dimension_OCHA}")






####### ** 12 **       ------------------------------  COSMESI ------------------------------------------     #######


# Define the colors
colors = {
    "light_beige": "FFF2CC",
    "light_orange": "F4B183",
    "dark_orange": "ED7D31",
    "darker_orange": "C65911",
    "light_blue": "DDEBF7",
    "light_pink": "b3b389",
    "light_yellow": "ffffc5",
    "white": "FFFFFF",
    "bluepin": "004bb4",
    'gray': 'e0e0e0'
}

# Define the columns to color
color_mapping = {
    label_perc2: colors["light_beige"],
    label_tot2: colors["light_beige"],
    label_perc3: colors["light_orange"],
    label_tot3: colors["light_orange"],
    label_perc4: colors["dark_orange"],
    label_tot4: colors["dark_orange"],
    label_perc5: colors["darker_orange"],
    label_tot5: colors["darker_orange"],
    label_perc_tot: colors["light_blue"],
    label_admin_severity: colors["light_blue"],
    label_tot: colors["light_blue"]
}


# Define the colors
colors_dimension = {
    "light_beige": "ebecc7",
    "light_orange": "c7ebec",
    "dark_orange": "c7d9ec",
    "darker_orange": "c7ecdb",
    'darker2_orange':'b3d3d4',
    "light_blue": "DDEBF7",
    "light_pink": "b3b389",
    "light_yellow": "ffffc5",
    "white": "FFFFFF",
    "bluepin": "004bb4",
    'gray': 'e0e0e0'
}

# Define the columns to color
color_mapping_dimension = {
    label_perc_out: colors_dimension["light_beige"],
    label_tot_out: colors_dimension["light_beige"],
    label_perc_acc: colors_dimension["light_orange"],
    label_tot_acc: colors_dimension["light_orange"],
    label_perc_agg: colors_dimension["dark_orange"],
    label_tot_agg: colors_dimension["dark_orange"],
    label_perc_lc: colors_dimension["darker_orange"],
    label_tot_lc: colors_dimension["darker_orange"],
    label_perc_penv: colors_dimension["darker2_orange"],
    label_tot_penv: colors_dimension["darker2_orange"],
    label_dimension_perc_tot: colors_dimension["light_blue"],
    label_dimension_tot: colors_dimension["light_blue"]
}

# List of columns that require specific alignment
alignment_columns = list(color_mapping.keys())



# Apply formatting to both output files
apply_formatting(output_JIAF, color_mapping, alignment_columns)
apply_formatting(output_OCHA, color_mapping, alignment_columns)
# Apply formatting to both output files
apply_formatting_dimension(output_dimension_JIAF, color_mapping_dimension, alignment_columns)
apply_formatting_dimension(output_dimension_OCHA, color_mapping_dimension, alignment_columns)



print(f"Formatting applied and files saved to {output_JIAF} and {output_OCHA}.")
print(f"Formatting applied and files saved to {output_dimension_JIAF} and {output_dimension_OCHA}.")

## saving excel 






















file_path = 'output_test/edu_data_filtered_test.xlsx'

# Save the DataFrame to an Excel file
edu_data.to_excel(file_path, index=False, engine='openpyxl')



output_file_path_test_strata_gender = 'output_test/severity_with_additional_strata_gender.xlsx'
output_file_path_test_strata_cycle = 'output_test/severity_with_additional_strata_cycle.xlsx'

with pd.ExcelWriter(output_file_path_test_strata_gender, engine='openpyxl') as writer:
    for group_name, df_group in severity_by_gender_list.items():
        # Each DataFrame is written to a separate sheet named after the group
        df_group.to_excel(writer, sheet_name=str(group_name), index=False)
with pd.ExcelWriter(output_file_path_test_strata_cycle, engine='openpyxl') as writer:
    for group_name, df_group in severity_by_cycle_list.items():
        # Each DataFrame is written to a separate sheet named after the group
        df_group.to_excel(writer, sheet_name=str(group_name), index=False)