import numpy as np
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.cell import MergedCell  # Import MergedCell
from openpyxl.utils.dataframe import dataframe_to_rows


int_2 = '2.0'
int_3 = '3.0'
int_4 = '4.0'
int_5 = '5.0'
label_perc2 = '% 1-2'
label_perc3 = '% 3'
label_perc4 = '% 4'
label_perc5 = '% 5'
label_tot2 = '# 1-2'
label_tot3 = '# 3'
label_tot4 = '# 4'
label_tot5 = '# 5'
label_perc_tot = '% Tot PiN (3+)'
label_tot = '# Tot PiN (3+)'
label_admin_severity = 'Area severity'
label_tot_population = 'TotN'

int_acc = 'access'
int_agg= 'aggravating circumstances'
int_lc = 'learning condition'
int_penv = 'protected environment'
int_out = 'not falling within the PiN dimensions'
label_perc_acc = '% Access'
label_perc_agg= '% Aggravating circumstances'
label_perc_lc = '% Learning conditions'
label_perc_penv = '% Protected environment'
label_perc_out = '% Not falling within the PiN dimensions'
label_tot_acc = '# Access'
label_tot_agg= '# Aggravating circumstances'
label_tot_lc = '# Learning conditions'
label_tot_penv = '# Protected environment'
label_tot_out = '# Not falling within the PiN dimensions'

label_dimension_perc_tot = '% Tot in PiN Dimensions'
label_dimension_tot = '# Tot in PiN Dimensions'

label_dimension_tot_population = 'TotN'



# Define the colors
colors = {
    "light_beige": "FFF2CC",
    "light_orange": "F4B183",
    "dark_orange": "ED7D31",
    "darker_orange": "C65911",
    "light_blue": "DDEBF7",
    "light_pink": "b3b389",
    "light_yellow": "ffffc5",
    "white": "FFFFFF",
    "bluepin": "004bb4",
    'gray': 'e0e0e0',
    'stratagray': 'F0F0F0'
}
# Define the columns to color
color_mapping = {
    label_perc2: colors["light_beige"],
    label_tot2: colors["light_beige"],
    label_perc3: colors["light_orange"],
    label_tot3: colors["light_orange"],
    label_perc4: colors["dark_orange"],
    label_tot4: colors["dark_orange"],
    label_perc5: colors["darker_orange"],
    label_tot5: colors["darker_orange"],
    label_perc_tot: colors["light_blue"],
    label_admin_severity: colors["light_blue"],
    label_tot: colors["light_blue"]
}
# Define the colors
colors_dimension = {
    "light_beige": "ebecc7",
    "light_orange": "c7ebec",
    "dark_orange": "c7d9ec",
    "darker_orange": "c7ecdb",
    'darker2_orange':'b3d3d4',
    "light_blue": "DDEBF7",
    "light_pink": "b3b389",
    "light_yellow": "ffffc5",
    "white": "FFFFFF",
    "bluepin": "004bb4",
    'gray': 'e0e0e0',
    'stratagray': 'F0F0F0'
}
# Define the columns to color
color_mapping_dimension = {
    label_perc_out: colors_dimension["light_beige"],
    label_tot_out: colors_dimension["light_beige"],
    label_perc_acc: colors_dimension["light_orange"],
    label_tot_acc: colors_dimension["light_orange"],
    label_perc_agg: colors_dimension["dark_orange"],
    label_tot_agg: colors_dimension["dark_orange"],
    label_perc_lc: colors_dimension["darker_orange"],
    label_tot_lc: colors_dimension["darker_orange"],
    label_perc_penv: colors_dimension["darker2_orange"],
    label_tot_penv: colors_dimension["darker2_orange"],
    label_dimension_perc_tot: colors_dimension["light_blue"],
    label_dimension_tot: colors_dimension["light_blue"]
}

alignment_columns = list(color_mapping.keys())
def apply_final_formatting(workbook, overview_df, small_overview_df, admin_var):
    for ws in workbook.worksheets:
        


        if ws.title == "PiN TOTAL":
            # Clear existing content in the worksheet
            ws.delete_rows(1, ws.max_row)

            # Add empty rows at the top and adjust columns
            ws.insert_rows(1, 4)
            ws.insert_cols(1, 2)

            # Write small_overview_df on the left (columns C-E)
            for r_idx, row in enumerate(dataframe_to_rows(small_overview_df, index=False, header=True), start=5):
                for c_idx, value in enumerate(row, start=3):  # Start writing in column C
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Write overview_df on the right (columns H onwards)
            for r_idx, row in enumerate(dataframe_to_rows(overview_df, index=False, header=True), start=5):
                for c_idx, value in enumerate(row, start=8):  # Start writing in column H
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Merge the title cell across all filled columns and ensure it is centered
            max_col = ws.max_column
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=max_col)
            title_cell = ws.cell(row=1, column=3)
            title_cell.value = "PiN TOTAL"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Set header formatting (bold, orange fill)
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=colors["light_orange"], end_color=colors["light_orange"], fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply color formatting and borders based on the content of cells
            for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
                strata_value = row[2].value  # Adjusting for zero-indexing; column C is the 'Strata' column
                
                if strata_value == "TOTAL (5-17 y.o.)":
                    fill_color = colors["bluepin"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.font = Font(color=colors["white"], bold=True)  # Set text color to white and bold
                elif strata_value in ["Girls", "Boys"]:
                    fill_color = colors["gray"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                elif strata_value in ['Female', 'Male', 'ECE (5 y.o.)', 'Children with disability']:
                    fill_color = colors["stratagray"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                else:
                    fill_color = colors["light_blue"]
                    for cell in row:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Set column widths based on content
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if cell.value and not isinstance(cell, MergedCell):
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                                column_letter = cell.column_letter  # Get the column letter
                            except Exception as e:
                                print(f"Error: {e}, Cell: {cell}")

                    if column_letter:
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column_letter].width = adjusted_width

            # Set the first two columns (A and B) to white background
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")

            # Apply specific formatting to columns F and G
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=6, max_col=7):
                for cell in row:
                    # Example formatting, adjust as needed
                    cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")
                    cell.alignment = Alignment(horizontal='right', vertical='center')

               # Set background for empty cells and apply black borders to cells with values
            thin_border = Border(
                left=Side(style='thin', color="000000"),
                right=Side(style='thin', color="000000"),
                top=Side(style='thin', color="000000"),
                bottom=Side(style='thin', color="000000")
            )

            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=ws.max_column):  # Adjust columns as needed
                for cell in row:
                    if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):  
                        cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")
                    else:  # If the cell has a value, apply black border
                        cell.border = thin_border

        if ws.title != "PiN TOTAL":
            ws.insert_rows(1, 4)
            # Add empty columns to the left
            ws.insert_cols(1, 4)
            # Add the sheet name as a title in the first row
            title = ws.title
            title += " (5-17 y.o.)"
            max_col = ws.max_column
            ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=max_col)
            title_cell = ws.cell(row=1, column=5)
            title_cell.value = title
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Set column widths based on content
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if cell.value and not isinstance(cell, MergedCell):
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                            column_letter = cell.column_letter  # Get the column letter
                        except Exception as e:
                            print(f"Error: {e}, Cell: {cell}")

                if column_letter:
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Bold specific columns
            for row in ws.iter_rows(min_row=6, max_col=ws.max_column, max_row=ws.max_row):  # Start from the data row
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value  # Row 5 contains the headers
                    if col_name in [label_perc_tot, label_admin_severity, label_tot, admin_var]:
                        cell.font = Font(bold=True)  # Apply bold font
        
            # Apply formatting to the headers
            for cell in ws[5]:  # Header row is now the 5th row
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply color to specific columns and make borders visible
            for row in ws.iter_rows(min_row=5, max_col=ws.max_column, max_row=ws.max_row):
                for cell in row:
                    col_name = ws.cell(row=5, column=cell.column).value
                    # Apply color based on mapping
                    if col_name in color_mapping:
                        cell.fill = PatternFill(start_color=color_mapping[col_name], end_color=color_mapping[col_name], fill_type="solid", patternType='solid')
                    # Apply alignment
                    if col_name in alignment_columns:
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                    # Apply border, but skip the first four columns
                    if cell.column > 4:
                        if row[0].row == 5:  # Bold top border for header
                            cell.border = Border(
                                top=Side(style="medium"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[0]:  # Bold left border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="medium"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                        elif cell == row[-1]:  # Bold right border for each row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="medium"),
                                bottom=Side(style="thin"),
                            )
                        elif row[0].row == ws.max_row:  # Bold bottom border for last row
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="medium"),
                            )
                        else:
                            cell.border = Border(
                                top=Side(style="thin"),
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                bottom=Side(style="thin"),
                            )
                    else:
                        cell.fill = PatternFill(start_color=colors["white"], end_color=colors["white"], fill_type="solid")  # White background for the first four columns
                        cell.border = None  # No border for the first four columns



    return workbook


# Function to create output with final formatting
def create_output(dataframes, overview_df, small_overview_df, overview_sheet_name, admin_var, ocha=True, tot_severity=None):
    output = BytesIO()
    with pd.ExcelWriter(output) as writer:
        # Only write the overview sheet if ocha is True
        if ocha:
            overview_df.to_excel(writer, sheet_name=overview_sheet_name, index=False)

        # Write the tot_severity sheet if it is provided
        if tot_severity is not None:
            tot_severity.to_excel(writer, sheet_name='Overall PiN and severity', index=False)

        # Write the category sheets
        for category, df in dataframes.items():
            sheet_name = f"{overview_sheet_name.split()[0]} -- {category}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    output.seek(0)
    workbook = load_workbook(output)

    # Apply the final formatting to the workbook
    workbook = apply_final_formatting(workbook, overview_df, small_overview_df, admin_var)
    
    formatted_output = BytesIO()
    workbook.save(formatted_output)
    formatted_output.seek(0)

    return formatted_output



